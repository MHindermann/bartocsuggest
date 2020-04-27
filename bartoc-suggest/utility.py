""" utility.py

Utility functions. """

from __future__ import annotations
from typing import Optional, Dict
from os import path
from openpyxl import load_workbook
from jskos import Concept, ConceptScheme, LanguageMap

import json

DIR = path.dirname(path.abspath(__file__))


class Utility:
    """ Utility functions """

    @classmethod
    def load_file(cls, filename: str) -> Optional[ConceptScheme]:
        """ Load a file.

        filename: MUST use complete file path. """

        # stop if file does not exist
        ###!!!! filename = path.join(DIR, "input/" + filename)
        if path.exists(filename) is False:
            print(f"ERROR: File {filename} does not exist!")
            return None

        # choose method depending on file type:
        if filename.endswith(".xlsx"):
            workbook = load_workbook(filename)
            return cls.xlsx2jskos(workbook)
        elif filename.endswith(".json"):
            # TODO: add JSON support
            pass
        else:
            pass

    @classmethod
    def xlsx2jskos(cls, workbook) -> ConceptScheme:
        """ Transform XLSX to JSKOS """

        scheme = ConceptScheme()
        for worksheet in workbook:
            for row in worksheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
                if row[0] is None:
                    continue
                else:
                    concept = Concept(preflabel=LanguageMap({"en": row[0]}))  # TODO: automate language detection
                    scheme.concepts.append(concept)

        return scheme

    @classmethod
    def list2jskos(cls, input_list: list) -> ConceptScheme:
        """ Transform list to JSKOS """

        scheme = ConceptScheme()
        for item in input_list:
            concept = Concept(preflabel=LanguageMap({"en": item}))  # TODO: automate language detection
            scheme.concepts.append(concept)

        return scheme

    @classmethod
    def save_json(cls, json_object: Dict, preload_folder: str, number: int):
        """ Save a JSON object to a file.

        preload_folder: MUST use complete folder path. """

        filename = preload_folder + f"query_{number}.json"

        with open(filename, "w") as file:
            json.dump(json_object, file)

        print(f"{filename} preloaded")

    @classmethod
    def load_json(cls, preload_folder: str, number: int) -> Dict:
        """ Load a JSON object as Python dictionary from a file """

        filename = preload_folder + f"query_{number}.json"

        with open(filename) as file:

            json_object = json.load(file)

            return json_object
