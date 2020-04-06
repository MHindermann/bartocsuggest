""" main.py """

from __future__ import annotations
from typing import Union, Optional, Set, List

import requests

""" outline of modules
----------------------
MODULE 1: read input data from local file
MODULE 2: send request to FAST
MODULE 3: analyse result
MODULE 4: output result as x
"""

"""
MODULE 1:

1. define which files can be used and what format is required; provide template
2. json, excel should be fine for starters
"""

from os import path
from openpyxl import load_workbook

DIR = path.dirname(path.abspath(__file__))


class Utility:
    """ Utility functions """

    @classmethod
    def load_file(cls, filename) -> Optional[ConceptScheme]:
        """ Load a file from /input """

        # stop if file does not exist
        filename = path.join(DIR, "input/" + filename)
        if path.exists(filename) is False:
            print(f"ERROR: File {filename} does not exist!")
            return None

        # choose method depending on file type:
        if filename.endswith(".xlsx"):
            workbook = load_workbook(filename)
            return cls.xlsx2jskos(workbook)
        elif filename.endswith(".json"):
            pass
        else:
            pass

    @classmethod
    def xlsx2jskos(cls, workbook) -> ConceptScheme:
        """ Transform XLSX from template to JSKOS """

        scheme = ConceptScheme()
        for worksheet in workbook:
            for row in worksheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
                if row is None:
                    continue
                else:
                    concept = Concept(preflabel=LanguageMap({"en": row[0]}))  # TODO: automate language detection
                    scheme.concepts.add(concept)
        return scheme

def parse(file):
    """ """
    data = open(file)


Utility.load_file("owcm_index.xlsx")

# TODO: create concept scheme from input file


""" BARTOC FAST query module

This module sends queries to BARTOC FAST API and retrieves the results.
"""

FAST_API = "https://bartoc-fast.ub.unibas.ch/bartocfast/api?"

FAST_SOURCES = ["LuSTRE",
                "GettyULAN",
                "GettyAAT",
                "GettyTGN",
                "AgroVoc",
                "Bartoc",
                "data.ub.uio.no",
                "Finto",
                "Legilux",
                "MIMO",
                "UNESCO",
                "OZCAR-Theia",
                "GACS",
                "51.15.194.251",
                "UAAV",
                "HTW-Chur",
                "FORTH",
                "Irstea",
                "DemoVoc",
                "ScoLOMFR",
                "lobid-gnd",
                "Research-Vocabularies-Australia",
                "Loterre"]


class Query:
    """ A BARTOC FAST query """

    def __init__(self, searchword: str, maxsearchtime: int = 5, disabled: List[str] = None) -> None:
        self.searchword = searchword
        self.maxsearchtime = maxsearchtime
        if disabled is None:
            self.disabled = ["Research-Vocabularies-Australia", "Loterre"]

    def get(self) -> requests.models.Response:
        """ Send query as HTTP request to BARTOC FAST API """

        payload = {"searchword": self.searchword,
                   "maxsearchtime": self.maxsearchtime,
                   "disabled": self.disabled}

        return requests.get(url=FAST_API, params=payload)


""" JSKOS classes 

Built based on  https://gbv.github.io/jskos/context.json """


class LanguageMap:
    """ http://gbv.github.io/jskos/jskos.html#language-map """

    def __init__(self, _mapping: dict) -> None:
        self._mapping = _mapping

    def add(self, label: str, language: str):
        """ Add a value for a language"""

        self._mapping.update({label: language})

    def get_value(self, language: str) -> Optional[str]:
        """ Get the value of a language if any """

        return self._mapping.get(language)


class Resource:
    """ http://gbv.github.io/jskos/jskos.html#resource """

    def __init__(self,
                 uri: str = None,
                 form: Set[str] = None,  # form = type
                 context: str = None
                 ) -> None:
        self.uri = uri
        self.form = form
        if context is None:
            self.context = "https://gbv.github.io/jskos/context.json"


class Item(Resource):
    """ http://gbv.github.io/jskos/jskos.html#item """

    def __init__(self,
                 uri: str = None,
                 form: Set[str] = None,
                 context: str = None,
                 url: str = None,
                 preflabel: LanguageMap = None
                 ) -> None:
        if url is None:
            self.url = uri
        self.preflabel = preflabel
        super().__init__(uri, form, context)


class Concept(Item):
    """ http://gbv.github.io/jskos/jskos.html#concept """

    def __init__(self,
                 uri: str = None,
                 form: Set[str] = None,
                 context: str = None,
                 url: str = None,
                 preflabel: LanguageMap = None,
                 inscheme: set = None
                 ) -> None:
        self.inscheme = inscheme
        super().__init__(uri, form, context, url, preflabel)


class ConceptScheme(Item):
    """ http://gbv.github.io/jskos/jskos.html#concept-schemes """

    def __init__(self,
                 uri: str = None,
                 form: Set[str] = None,
                 context: str = None,
                 url: str = None,
                 preflabel: LanguageMap = None,
                 concepts: Set[Concept] = None
                 ) -> None:
        self.concepts = concepts
        super().__init__(uri, form, context, url, preflabel)
