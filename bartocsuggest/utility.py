""" utility.py

Utility functions. """

from __future__ import annotations
from typing import Optional, Dict, List, Union
from os import path
from datetime import datetime
from openpyxl import load_workbook
from json import dump, dumps, load
from annif_client import AnnifClient

from .jskos import _Concept, _ConceptScheme, _LanguageMap


class _Utility:
    """ Utility functions """

    @classmethod
    def load_file(cls, filename: str, language: str = "und") -> Optional[_ConceptScheme]:
        """ Load a file.

        :param filename: the name of the file including its complete path
        :param language: the language of the words given as RFC 3066 language tag, defaults to "und"
        """

        # stop if file does not exist
        if path.exists(filename) is False:
            print(f"ERROR: File {filename} does not exist!")
            return None

        # choose method depending on file type:
        if filename.endswith(".xlsx"):
            workbook = load_workbook(filename)
            return cls.xlsx2jskos(workbook, language)
        elif filename.endswith(".json"):
            # TODO: add JSON support
            pass
        else:
            pass

    @classmethod
    def xlsx2jskos(cls, workbook, language: str = "und") -> _ConceptScheme:
        """ Transform a XLSX workbook into a JSKOS concept scheme.

        The XLSX workbook's data structure MUST be as follows: one column with one row per word.

        :param workbook: the XLSX workbook
        :param language: the language of the words given as RFC 3066 language tag, defaults to "und"
        """

        scheme = _ConceptScheme()
        for worksheet in workbook:
            for row in worksheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
                if row[0] is None:
                    continue
                else:
                    concept = _Concept(pref_label=_LanguageMap({language.lower(): row[0]}))
                    scheme.concepts.append(concept)

        return scheme

    @classmethod
    def list2jskos(cls, input_list: list, language: str = "und") -> _ConceptScheme:
        """ Transform the list into a JSKOS concept scheme.

        :param input_list: the list of words
        :param language: the language of the words given as RFC 3066 language tag, defaults to "und"
        """

        scheme = _ConceptScheme()
        for item in input_list:
            concept = _Concept(pref_label=_LanguageMap({language.lower(): item}))
            scheme.concepts.append(concept)

        return scheme

    @classmethod
    def annif2jskos(cls, annif_suggestion: List[dict], annif_project_id: str) -> _ConceptScheme:
        """ Transform an Annif suggestion to JSKOS.

        :param annif_suggestion: the output of calling AnnifClient.suggest
        :param annif_project_id: Annif API project identifier
        """

        # get project details:
        annif = AnnifClient()
        project = annif.get_project(annif_project_id)
        language = project.get("language")
        name = project.get("name")

        # setup concept scheme based on project details:
        # TODO: add project uri
        scheme = _ConceptScheme(pref_label=_LanguageMap({language: name}))

        # make concept from result and add to concept scheme:
        for result in annif_suggestion:
            label = result.get("label")
            uri = result.get("uri")
            # TODO: add inscheme = concept scheme's uri
            concept = _Concept(uri=uri, pref_label=_LanguageMap({language: label}))
            scheme.concepts.append(concept)

        return scheme

    @classmethod
    def save_json(cls, dictionary: Dict, folder: str, filename: str = None):
        """ Save a dictionary as JSON file.

        :param dictionary: the dictionary to be saved
        :param folder: the folder to write the JSON file in (MUST use complete folder path)
        :param filename: the name of the file, defaults to None
        """

        if filename is None:
            filename = str(datetime.now()).split(".")[0].replace(":", "-")

        full_filename = folder + f"{filename}.json"
        with open(full_filename, "w") as file:
            dump(dictionary, file)

    @classmethod
    def load_json(cls, folder: str, filename: str) -> Dict:
        """ Load a JSON object as dictionary from a file """

        filename = folder + f"{filename}.json"

        with open(filename) as file:
            dictionary = load(file)

            return dictionary

    @classmethod
    def print_json(cls, dictionary: dict, indent: int = 2) -> None:
        """ Print a dictionary as JSON to the console.

        :param dictionary: the dictionary to be printed
        :param indent: indentation for pretty printing, defaults to 2
        """

        print(dumps(dictionary, indent=indent))
